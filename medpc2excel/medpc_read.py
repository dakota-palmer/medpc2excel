import re  #for word processing
import pandas as pd    #for data loading and manipulation
import os #for access folder
import numpy as np   #for calculaiton
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict
Tree= lambda: defaultdict(Tree)

def medpc_read (file, working_var_label='', rat_id=None, save=True, skipold = True, override = True, replace = True, log=''):
    '''
    Inputs:
    1. file (str, path)
    2. working_var_label (str, default is '')
    3. rat_id   (array like)
    4. save (Bolean value,default is True)
    5. override (Bolean value, default is True)
    6. replace  (Bolean value, default is True)
    
    Outputs:
    1. TS_df_tree (a tree, like {'date':{'rat':df}})
    2. log (string, capture essential events)
    
    Example:
    TS_df_tree, log =  medpc2excel (r'C:\20200229_test', override = True)
    
    '''
    alldata_tree = Tree()
    subject_list= []
    MSN_dict={}
    nowtime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    progamnames = [] #programnames = will be a list of MPC 'programs' (compiled in loop through datasets)
    
    #################################
    #load raw file into dict
    #################################
    with open(file, 'r') as f:
        datasets = f.read().split('Start Date: ') #Split the data file at Start Dates
        TS_var_name_maps = {}
        arrayA_name_maps = {}
        for n in range(1, len(datasets)): # Iterate through all subjects' data in the file
            theData = datasets[n]
            thisDate = (datetime.strptime(theData[0:8], "%m/%d/%y")).strftime("%Y-%m-%d")
            thisDate = thisDate.replace('-','')
            if theData.find('\r\n'):
                theData = theData.replace('\r\n', '\n')
                splitOffData = theData.split('MSN:')
                match = re.search('Subject: .*\n', splitOffData[0])
                subject = match.group(0).split(':')[1].strip()
                subject_list.append(subject)
                variables = splitOffData[1].split('\n')
                match2 = re.search('Box: .*\n', splitOffData[0])
                box = match2.group(0).split(':')[1].strip()
                programname = variables.pop(0).strip()
                if programname not in progamnames:
                    progamnames.append(programname)
            #if there is date needed to be remove from a ID's data
            #then skip all the extraction process below
            #only extract data that we need
            if rat_id == None:
                pass
            else:
                rat_id = [str(i) for i in rat_id] #force convert the value type inside the list
                if str(subject) not in rat_id:
                    continue
            
            #recognize all defined variable name in MSN protocol
            MSN_folder = os.path.dirname(file)
            MSN = os.path.join(MSN_folder, programname+'.MPC') #find the MPC matching that listed in this data file (theData)  
            try:
                TS_var_name_map = {}
                arrayA_name_map = {}
                with open(MSN, 'r') as f:
                    for n, line in enumerate(f):
                        if 'DIM' in line:
                            pat = re.compile(r'(DIM\s*)(\w)([\s=\d]*)([\s\\]*)(\w*\s*\w*)(\s+)([\w\(\)]*)')
                            var, name = pat.search(line).group(2), pat.search(line).group(5)
                            if var != working_var_label: 
                                if name != '':
                                    name = re.sub('[\s]*', '', name)
                                    TS_var_name_map[var] = "(%s)"%var+name
                                    TS_var_name_maps[programname] = TS_var_name_map
                        elif working_var_label != '':
                            if re.match(r'\s*\\\s*%s\(\d*\)'%working_var_label, line):
                                pat = re.compile(r'(\s*\\\s*)(%s\()(\d*)(\))\W*([\w\s\(\)]*)([\w\s\(\),\/]*)'%working_var_label) 
                                idx, name = pat.search(line).group(3), pat.search(line).group(5)
                                arrayA_name_map[idx]='%s(%s)'%(working_var_label,idx)+name.strip('\n')
                                arrayA_name_maps[programname] = arrayA_name_map
            
            except:
                log += nowtime+'>>\t'+'Error!Pleas provide MSN programe in the data folder: %s'%MSN +'\n'
            
            varnameLists =  [line for line in variables if not re.search('\w:\s+', line) and line != '']
            #find common variable names between MSN protocol and datafile
            common_vars = [var.strip(":") for var in varnameLists if var.strip(":") in TS_var_name_map.keys()]
            # #assume array A is alwasy store working variables
            # common_vars += ['A'] # remove this line cannot assume there is always A array inside the file
            common_vars.sort()  #sort the common variables to make sure they are in order
            data_dict = {}
            
            for idx, var in enumerate(common_vars, 1):
                if idx < len(common_vars):
                    start = variables.index(common_vars[idx-1]+":")
                    end = variables.index(common_vars[idx]+":")
                    data = variables[start+1:end]
                else:
                    start = variables.index(common_vars[idx-1]+":")
                    data = variables[start+1:]
                temp = []
    
                for d in data:
                    if d!='':
                        temp += re.split('\s+',d.split(':')[1])
                        temp.remove('')
                    #convert str --> numbers
                data_dict[var.strip(':')] =  pd.to_numeric(pd.Series(temp, name = var.strip(':'),dtype = float))
            alldata_tree[programname][thisDate][subject] = data_dict
          
            #temporary metadata for this file in MSN_dict? to be compared and collated across all files in MSN_file?
            if thisDate not in MSN_dict.keys(): 
                MSN_dict[thisDate] = pd.DataFrame({'ID':[subject],'Box':[box],'MSN':[programname]})
            else:
                MSN_dict[thisDate] = MSN_dict[thisDate].append({'ID':subject,'Box':box,'MSN':programname},ignore_index=True)
    
    ####################################
    # Get timestamps for each variables
    ####################################
    TS_df_tree = Tree() 
    for program_nm in progamnames:
        for d, alldata_dict in alldata_tree[program_nm].items():
            temp2_dfs = [] # for working vars dataframe
            for rat, data_d in alldata_dict.items():
                temp_df_list = []
                for var, nm in TS_var_name_maps[program_nm].items():
                    data_d[var].name = nm
                    temp_df_list.append(data_d[var])     
                TS_df_tree[d][rat] = pd.concat(temp_df_list,axis = 1)
    
    ####################################
    # Get values for working variables
    ####################################
    if working_var_label != '':
        workingVar_tree = Tree() #computed, but not return for this version
        for program_nm in progamnames:
            for d, alldata_dict in alldata_tree[program_nm].items():
                temp2_dfs = [] # for working vars dataframe
                for rat, data_d in alldata_dict.items():
                    temp2 = {}
                    for k, nm in arrayA_name_maps[program_nm].items():
                        try:
                            temp2[nm] = data_d[working_var_label].iloc[int(k)]
                        except IndexError:
                            log += nowtime+'>>\t'+'Error!%s in MSN (%s) cannnot find in file %s'%(nm,programname,file) + '\n'
                            temp2[nm] = np.nan
                    temp2_df = pd.DataFrame.from_dict(temp2, orient = 'index', columns = [rat])
                    temp2_dfs.append(temp2_df)
                    temp2_dfs_concat = pd.concat(temp2_dfs, axis = 1)
                workingVar_tree[d][program_nm] = temp2_dfs_concat
        
        workingVar_dfs = {}
        for d, v in workingVar_tree.items():
            workingVar_dfs[d] =  pd.concat(v.values(), axis = 1, join = 'outer')
            workingVar_dfs[d].sort_index(ascending=True, axis=1, inplace = True)
    else:
        workingVar_dfs = {}
    
    ######################################
    # Save data into local files
    ######################################
    if save:
        file_path = os.path.dirname(file)
        for d, TS_df_dict in TS_df_tree.items():
            filename = os.path.join(file_path, '%s.xlsx'%d)
            
            if os.path.exists(filename): #if the file has exists
                
                if not skipold: # use doesn't want to skip the old file 
                    #first exame whether it contains the subject above
                    x1 = pd.ExcelFile(filename)
                    
                    #below reads only single msn sheet as df and throws error 
                    # x1 = pd.read_excel(filename, engine='openpyxl') #dp 2022-07-01 update for newer pandas
                    # x1 = pd.read_excel(filename)#, engine='openpyxl') #dp 2022-07-01 update
                    # x1 = pd.read_csv(filename) #dp 2022-07-01 update

                    #look for this current file subj sheet in xlsx
                    overlap = list(set(x1.sheet_names) & set(TS_df_dict.keys())) #search for existing sheet for this subject?
                    
                    # If user want to complete overide the old file into a new one
                    if override:
                        with pd.ExcelWriter(filename, mode = 'r', engine='openpyxl') as writer: # pylint: disable=abstract-class-instantiated
                            MSN_dict[d].to_excel(writer,sheet_name='MSNs',index=False)
                            if working_var_label !='':
                                workingVar_dfs[d].to_excel(writer, sheet_name='Summary_(%s)'%working_var_label, index = True)
                            for sheet, df in TS_df_dict.items():
                                df.to_excel(writer, sheet_name = sheet, index=False)
                            log += nowtime+'>>\t'+'Override the exisitng local excel file %s'%filename +'\n'
                        
                    else: # No override
                        # If user want to replace
                        if replace:
                            if len(overlap)> 0: # replace existing data and concat new data into files
                                
                                book = load_workbook(filename)
                                writer = pd.ExcelWriter(filename, engine = 'openpyxl')
                                # writer = pd.ExcelWriter(filename, mode='a', engine = 'openpyxl') #dp 2022-07-01 need mode='a' ?

                                writer.book = book
                                MSNs_file = pd.read_excel(filename,sheet_name = 'MSNs')
                                MSNs_file['ID'] = MSNs_file['ID'].astype('str')
                                if working_var_label !='':
                                    summary_file = pd.read_excel(filename, sheet_name = 'Summary_(%s)'%working_var_label)
                                    summary_file.set_index('Unnamed: 0', inplace = True)
                                    summary_file.columns = summary_file.columns.astype(str)
                                
                                new = list(set(TS_df_dict.keys())^set(x1.sheet_names))
                                new = [n for n in new if n!= 'Summary_(%s)'%working_var_label and n!= 'MSNs']
                                if len(new)>0: # determine if this is partial overlap, if yes, only replace the overlapped part, keep the other original data and append new data.
                                
                                    new_MSNs_file = pd.concat([MSN_dict[d], MSNs_file[MSNs_file['ID'].isin(new)]], axis = 0) # Concat new MSNs info with old one, which doesn't contain the overlapped ID
                                    
                                    if working_var_label !='':
                                        working_new_col = list(workingVar_dfs[d].columns[workingVar_dfs[d].columns.isin(new)])
                                        new_col = list(summary_file.columns[summary_file.columns.isin(new)])
                                        new_summary_file = pd.concat([workingVar_dfs[d][overlap], workingVar_dfs[d][working_new_col], summary_file[new_col]], axis = 1)
                                        
                                    new_TS_df_dict = TS_df_dict.copy()
                                    new_in_excel  = list(set(x1.sheet_names) & set(new))
                                    for new_id in new_in_excel:
                                        new_TS_df_dict[new_id] = pd.read_excel(filename, sheet_name = new_id)
                                
                                else: # if this is a complete overlap. Then directly replace everything
                                    new_MSNs_file = MSN_dict[d].copy()
                                    if working_var_label !='':
                                        new_summary_file = workingVar_dfs[d].copy()
                                    new_TS_df_dict = TS_df_dict.copy()
                                
                                new_MSNs_file.sort_values(by = ['ID']) # sort ID from small to largest
                                new_summary_file.sort_index(ascending=True, axis=1, inplace = True) # sort column name
                                
                                with pd.ExcelWriter(filename, mode = 'r', engine='openpyxl') as writer: # pylint: disable=abstract-class-instantiated
                                    new_MSNs_file.to_excel(writer, sheet_name = 'MSNs', index = False)
                                    if working_var_label !='':
                                        new_summary_file.to_excel(writer, sheet_name = 'Summary_(%s)'%working_var_label, index = True)
                                    for sheet, df in new_TS_df_dict.items():
                                        df.to_excel(writer, sheet_name = sheet, index=False)
                                    log += nowtime+'>>\t'+'Replace overlapped MED-PC data in the existing local excel file %s'%filename +'\n'
                            
                            else: # no overlap, then append new data to the existing file
                                new = list(set(TS_df_dict.keys())^set(overlap))
                                MSNs_file = pd.read_excel(filename,sheet_name = 'MSNs')
                                
                                if MSNs_file.values.shape == MSN_dict[d].values.shape: #dp- compare the MSN_dict metadata values with those saved thus far in the .xlsx
                                    if np.prod(np.equal(MSNs_file.astype(str).values,MSN_dict[d].astype(str).values)):
                                        MSN_same = True
                                    else:
                                        MSN_same = False
                                else:    
                                    MSN_same = False
                                
                                if working_var_label !='':
                                    summary_file = pd.read_excel(filename, sheet_name = 'Summary_(%s)'%working_var_label)
                                    if summary_file.values.shape == workingVar_dfs[d].values.shape:
                                        if np.prod(np.equal(summary_file.astype(str).values, workingVar_dfs[d].atype(str).values)):
                                            if list(summary_file.columns) == list(workingVar_dfs[d].columns):
                                                summary_same = True
                                            else:
                                                summary_same = False
                                        else:
                                            summary_same = False
                                    else:
                                        summary_same = False
                                    
                                if not MSN_same or not summary_same: 
                                    book = load_workbook(filename)
                                    # writer = pd.ExcelWriter(filename, engine = 'openpyxl') #dp 2022-07-01 should have mode='a'?
                                    writer = pd.ExcelWriter(filename, mode='a', engine = 'openpyxl') #dp 2022-07-01 should have mode='a'?

                                    writer.book = book
                                    
                                    if not MSN_same:
                                        #this corrupts the excel file dp
                                        writer.sheets = {ws.title: ws for ws in book.worksheets if ws.title=='MSNs'} #get MSNs sheet from existing excel
                                        MSN_dict[d].to_excel(writer, sheet_name = 'MSNs', startrow = writer.sheets['MSNs'].max_row, header = None, index= False)
                                    
                                    if working_var_label !='':
                                        if not summary_same:
                                            # creat a new file and replace the old one
                                            std= book['Summary_(%s)'%working_var_label]
                                            book.remove(std)
                                            summary_file.set_index('Unnamed: 0', inplace = True)
                                            newfile = pd.concat([summary_file, workingVar_dfs[d]], axis = 1, join = 'outer')
                                            newfile.to_excel(writer, sheet_name = 'Summary_(%s)'%working_var_label, index = True) 
                                        writer.save()
                                        writer.close()
                                        
                                with pd.ExcelWriter(filename, mode = 'a', engine='openpyxl') as writer: # pylint: disable=abstract-class-instantiated
                                    for sheet in new:
                                        TS_df_dict[sheet].to_excel(writer, sheet_name = sheet, index=False) 
                                    log += nowtime+'>>\t'+'No overlap. Append new MED-PC data to an existing local excel file %s.'%filename +'\n'
                            
                                
                        else: # Use don't want to replace, then append new data into old file
                            #else, append new subject into this excel
                            new = list(set(TS_df_dict.keys())^set(overlap))
                            #dp 2022-07-01 why exclude with this conditional?
                            #-----tru but code isnt running in debugger
                            # it is running once this if is commented out...
                            # if len(new) > 0: 
                            MSNs_file = pd.read_excel(filename,sheet_name = 'MSNs')
                            #first check if shape of msn same, then values (maybe just trying to be efficient?)
                            if MSNs_file.values.shape == MSN_dict[d].values.shape: #dp if two MSNs summary are not the same append current one to the file one
                                if np.prod(np.equal(MSNs_file.astype(str).values,MSN_dict[d].astype(str).values)):
                                    MSN_same = True
                                else:
                                    MSN_same = False #makes it here without if()
                            else:    
                                MSN_same = False
                            #-----was wrapped in above conditional on len(new)--
                            
                            #dp add summary_same definition even if working_var_label is blank
                            if working_var_label == '':
                                summary_same= False
                            
                            if working_var_label !='':
                                summary_file = pd.read_excel(filename, sheet_name = 'Summary_(%s)'%working_var_label)
                                if summary_file.values.shape == workingVar_dfs[d].values.shape:
                                    if np.prod(np.equal(summary_file.astype(str).values, workingVar_dfs[d].astype(str).values)):
                                        if list(summary_file.columns) == list(workingVar_dfs[d].columns):
                                            summary_same = True
                                        else:
                                            summary_same = False
                                    else:
                                        summary_same = False
                                else:
                                    summary_same = False
                                    
                                #dp unindent 1 level
                                #DP THIS shouldnt be running in above if()? should run if working_var_label is blank too
                                #dp this should be running on 7/7/22 test append
                            if not MSN_same or not summary_same: #if two MSNs summary are not the same append current one to the file one
                                book = load_workbook(filename)
                                # writer = pd.ExcelWriter(filename, engine = 'openpyxl') #2022-07-01 dp add mode='a'
                                # writer = pd.ExcelWriter(filename, mode='a', engine = 'openpyxl') #2022-07-01 dp add mode='a'
                                #try add if_sheet_exists
                                writer = pd.ExcelWriter(filename, mode='a', engine = 'openpyxl', if_sheet_exists='overlay') #2022-07-01 dp add mode='a'
                                #overlay should allow extending of MSN sheet?
                                # overlay mode didn't work, just going to overwrite?
                                # writer = pd.ExcelWriter(filename, mode='a', engine = 'openpyxl', if_sheet_exists='replace') #2022-07-01 dp add mode='a'

                                
                                #dp 2022-07-01 this corrupts xlsx without mode='a'?
                                writer.book = book
                                
                                if not MSN_same:
                                    writer.sheets = {ws.title: ws for ws in book.worksheets if ws.title=='MSNs'} #get MSNs sheet from existing excel
                                    
                                    #dp 2022-07-01 seems this is where the MSN sheet should be growing 
                                    # MSN_dict[d].to_excel(writer, sheet_name = 'MSNs', startrow = writer.sheets['MSNs'].max_row, header = None, index= False)
                                    # *** ValueError: Sheet 'MSNs' already exists and if_sheet_exists is set to 'error'.                     
                                    
                                    #trying with diff arguments
                                    # MSN_dict[d].to_excel(writer, sheet_name = 'MSNs', header = None, index= False)
                                   
                                    # MSN_dict[d].to_excel(writer, sheet_name = 'MSNs', index= False)

                                    
                                    #dp could try appending as df and then overwriting old sheet?
                                    MSNs_all= pd.concat([MSNs_file, MSN_dict[d]])
                                    
                                    # MSNs_all.to_excel(writer, sheet_name = 'MSNs', startrow = writer.sheets['MSNs'].max_row, header = None, index= False)
                                    # MSNs_all.to_excel(writer, sheet_name = 'MSNs', header = None, index= False)
                                    # MSNs_all.to_excel(writer, sheet_name = 'MSNs', index= False)

                                    #try with writer like below- THIS WORKS!
                                    with pd.ExcelWriter(filename, mode='a', engine = 'openpyxl', if_sheet_exists='overlay') as writer:
                                        MSN_dict[d].to_excel(writer, sheet_name = 'MSNs', startrow = writer.sheets['MSNs'].max_row, header = None, index= False)        
                                        
                                    if working_var_label !='':
                                        if not summary_same:
                                            # creat a new file and replace the old one
                                            std= book['Summary_(%s)'%working_var_label]
                                            book.remove(std)
                                            summary_file.set_index('Unnamed: 0', inplace = True)
                                            newfile = pd.concat([summary_file, workingVar_dfs[d]], axis = 1, join = 'outer')
                                            newfile.to_excel(writer, sheet_name = 'Summary_(%s)'%working_var_label, index = True)
                                        writer.save()
                                        writer.close()
                                        
                                        #dp 7/7/22 this works if run tho isnt running
                                # with pd.ExcelWriter(filename, mode = 'a', engine='openpyxl') as writer: # pylint: disable=abstract-class-instantiated
                                with pd.ExcelWriter(filename, mode = 'a', engine='openpyxl', if_sheet_exists='overlay') as writer: # pylint: disable=abstract-class-instantiated
                                    for sheet in new:
                                        TS_df_dict[sheet].to_excel(writer, sheet_name = sheet, index=False) 
                                    log += nowtime+'>>\t'+'Append new MED-PC data to an existing local excel file %s. Old data was not changed.'%filename +'\n'
                            
            else: # file doesn't exist
                with pd.ExcelWriter(filename, engine='openpyxl') as writer: # pylint: disable=abstract-class-instantiated
                #mode a wont work if file not exist
                # with pd.ExcelWriter(filename, mode='a', engine='openpyxl') as writer: # pylint: disable=abstract-class-instantiated
                    MSN_dict[d].to_excel(writer,sheet_name='MSNs',index=False)
                    if working_var_label !='':
                        workingVar_dfs[d].to_excel(writer, sheet_name='Summary_(%s)'%working_var_label, index = True)
                    for sheet, df in TS_df_dict.items():
                        df.to_excel(writer, sheet_name = sheet, index=False)
                    log += nowtime+'>>\t'+'Extract MED-PC data file to a new local excel file %s'%file_path+'\n'

    return TS_df_tree, log